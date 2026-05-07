"""主解析器 —— 文本/Excel → ImportData(headers + rows + suggested mapping)。

设计上**只保留两条路径**:
- BeeCount 自家格式 —— 跟 web/mobile export 严格对齐(11 列 + header 在 row 0
  + 表头本地化),round-trip 必须等价
- generic —— 所有其它来源(支付宝 / 微信 / 银行账单 / 用户清洗的 Excel),
  统一用一张 alias 表 + 列数一致性启发寻找 header

之前为 alipay / wechat 各开了 sniff 分叉,但「找 header + 找列名」逻辑跟
generic 完全等价,只是 alias 集稍微偏向某一来源。维护两份反而容易漏 alias
(比如用户文件叫"分类"但 alipay parser 只认"类别"→ 误判 alipay 又匹不到列)。
统一到 generic 一张全集 alias 表更稳。

`detect_source_format` 仍保留 — 但只区分 beecount vs generic,不再细分支付宝/微信。
前端不展示给用户(避免误判误导,见 ImportPage 注释)。
"""
from __future__ import annotations

import csv
import io
import logging
from typing import cast

from .parsers.beecount import BeeCountParser
from .parsers.generic import GenericParser
from .schema import (
    ImportData,
    ImportFieldMapping,
    ParsedRow,
    ParseWarning,
    SourceFormat,
)

logger = logging.getLogger(__name__)

_PARSERS = {
    "beecount": BeeCountParser(),
    "generic": GenericParser(),
}


def detect_source_format(raw_text: str) -> SourceFormat:
    """只区分 BeeCount 自家 vs generic。
    支付宝 / 微信 / 银行账单 / 用户清洗的 Excel 全归 generic。"""
    sample = raw_text[:5_000].lower()
    if BeeCountParser().sniff(sample):
        return "beecount"
    return "generic"


def parse_csv_text(
    *,
    raw_text: str,
    forced_source: SourceFormat | None = None,
) -> ImportData:
    """端到端 CSV / TSV 文本解析。"""
    cleaned = _strip_bom_and_normalize(raw_text)
    rows_2d = _parse_csv_rows(cleaned)
    return _build_import_data(rows_2d=rows_2d, forced_source=forced_source)


def parse_excel_bytes(
    *,
    payload: bytes,
    forced_source: SourceFormat | None = None,
) -> ImportData:
    """端到端 .xlsx 解析 —— 用 openpyxl 读出 row[][] 后跟 CSV 走同一条路径。

    单 sheet only(支付宝 / 微信 / 用户自家 Excel 都是单 sheet 流水)。多
    sheet 文件取第一个,后续 sheet 静默忽略。
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for .xlsx import; pip install openpyxl"
        ) from exc

    wb = load_workbook(filename=io.BytesIO(payload), read_only=True, data_only=True)
    if not wb.sheetnames:
        return _build_import_data(rows_2d=[], forced_source=forced_source)
    ws = wb[wb.sheetnames[0]]

    rows_2d: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells: list[str] = []
        for v in row:
            if v is None:
                cells.append("")
            elif isinstance(v, (int, float)):
                # 整数 / 浮点直接 str,避免 1.0 / 4.5e3 等失真
                cells.append(_format_number(v))
            else:
                cells.append(str(v))
        rows_2d.append(cells)
    wb.close()

    return _build_import_data(rows_2d=rows_2d, forced_source=forced_source)


def _format_number(v: int | float) -> str:
    """openpyxl 把所有数字读成 float;尽量保持原样,避免 35 → 35.0 噪声。"""
    if isinstance(v, int):
        return str(v)
    if v == int(v):
        return str(int(v))
    return f"{v}"


def _build_import_data(
    *, rows_2d: list[list[str]], forced_source: SourceFormat | None
) -> ImportData:
    if not rows_2d:
        return ImportData(
            source_format="generic",
            headers=[],
            rows=[],
            suggested_mapping=ImportFieldMapping(),
            parse_warnings=[],
        )

    # 只用 BeeCount 嗅探;否则一律走 generic
    if forced_source is not None:
        source = forced_source
    else:
        sample_lower = "\n".join(",".join(r) for r in rows_2d[:30]).lower()
        source = "beecount" if BeeCountParser().sniff(sample_lower) else "generic"

    parser = _PARSERS[source]

    header_index = parser.find_header_row(rows_2d)
    if header_index < 0 or header_index >= len(rows_2d):
        logger.info(
            "import.parse no header row found by parser=%s, fallback header=0",
            source,
        )
        header_index = 0

    headers_raw = rows_2d[header_index]
    headers = [str(h).strip() for h in headers_raw]
    data_rows_2d = rows_2d[header_index + 1:]

    parsed_rows: list[ParsedRow] = []
    warnings: list[ParseWarning] = []
    expected_cols = len(headers)

    for offset, raw_cells in enumerate(data_rows_2d):
        row_number = header_index + 2 + offset
        if not raw_cells or all((c or "").strip() == "" for c in raw_cells):
            continue
        if len(raw_cells) != expected_cols:
            warnings.append(
                ParseWarning(
                    code="COLUMN_COUNT_MISMATCH",
                    row_number=row_number,
                    message=f"got {len(raw_cells)} columns, header has {expected_cols}",
                    raw_line=",".join(str(c) for c in raw_cells),
                )
            )
            cells = list(raw_cells)
            if len(cells) < expected_cols:
                cells.extend([""] * (expected_cols - len(cells)))
            else:
                cells = cells[:expected_cols]
        else:
            cells = list(raw_cells)
        cell_dict = {h: str(v) if v is not None else "" for h, v in zip(headers, cells)}
        parsed_rows.append(
            ParsedRow(
                row_number=row_number,
                cells=cell_dict,
                raw_line=",".join(str(c) for c in raw_cells),
            )
        )

    suggested = suggest_mapping(headers=headers, source=source)
    return ImportData(
        source_format=cast(SourceFormat, source),
        headers=headers,
        rows=parsed_rows,
        suggested_mapping=suggested,
        parse_warnings=warnings,
    )


def suggest_mapping(*, headers: list[str], source: SourceFormat) -> ImportFieldMapping:
    """每个 parser 自己定义最佳 mapping。仅 beecount / generic 两路。"""
    parser_key = source if source in _PARSERS else "generic"
    parser = _PARSERS[parser_key]
    return parser.suggest_mapping(headers)


# ──────────── 内部 helpers ────────────


_BOM = "\ufeff"


def _strip_bom_and_normalize(text: str) -> str:
    if text.startswith(_BOM):
        text = text[len(_BOM):]
    return text.replace("\r\n", "\n").replace("\r", "\n")


def _parse_csv_rows(text: str) -> list[list[str]]:
    if not text.strip():
        return []
    sample = text[:4096]
    dialect = None
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = None
    reader = csv.reader(
        io.StringIO(text),
        dialect=dialect or csv.excel,
    )
    return [row for row in reader]
